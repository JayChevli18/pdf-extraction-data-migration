"""Spreadsheet storage adapter."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from profile_backend.src.profile_backend.domain.models import ProfileRecord, headers_for_sheet


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"
    for col, header in enumerate(headers_for_sheet(), start=1):
        ws.cell(row=1, column=col, value=header)
    wb.save(path)


def append_record(path: Path, record: ProfileRecord) -> int:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    next_row = ws.max_row + 1
    values = record.to_row_list()
    for col, value in enumerate(values, start=1):
        ws.cell(row=next_row, column=col, value=value)
    wb.save(path)
    return next_row
