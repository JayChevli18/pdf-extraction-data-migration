"""Append rows to Excel without overwriting existing data."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook

from profile_backend.models import ProfileRecord, headers_for_sheet

logger = logging.getLogger("profile_backend.spreadsheet")


def ensure_workbook(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"
    for col, h in enumerate(headers_for_sheet(), start=1):
        ws.cell(row=1, column=col, value=h)
    wb.save(path)
    logger.info("Created spreadsheet %s", path)


def append_record(path: Path, record: ProfileRecord) -> int:
    """
    Append one row at the next empty line. Returns 1-based row number written.
    """
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    next_row = ws.max_row + 1
    values = record.to_row_list()
    for col, val in enumerate(values, start=1):
        ws.cell(row=next_row, column=col, value=val)
    wb.save(path)
    logger.info("Appended record at row %s to %s", next_row, path)
    return next_row
